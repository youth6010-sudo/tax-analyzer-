#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""검토표.xlsx → review-grid.json (셀 값·색·병합 그대로)."""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from datetime import date, datetime
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 필요: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

THEME_ORDER = (
    "lt1",
    "dk1",
    "lt2",
    "dk2",
    "accent1",
    "accent2",
    "accent3",
    "accent4",
    "accent5",
    "accent6",
    "hlink",
    "folHlink",
)
WHITE = "#FFFFFF"
BLACK = "#000000"

STAFF_NAMES = ("페리", "블루", "다야", "윈터", "리아")
CORP_BLOCKS = [
    ("블루", 1, 7),
    ("다야", 8, 14),
    ("윈터", 15, 21),
    ("리아", 22, 28),
    ("페리", 29, 35),
]

SECTION_LABELS = {
    "bookkeeping": "기장",
    "sincere": "성실",
    "corp_client": "업체",
    "agent": "신고",
    "transfer": "이관/폐업",
    "consult": "상담",
}


INCOME_MARKER_COL = 2
TRANSFER_KEYWORDS = ("폐업", "이관", "해지", "퇴사", "종료", "폐쇄", "말소", "중단")


def classify_section_marker(value) -> str | None:
    if value is None or not isinstance(value, str):
        return None
    raw = value.strip()
    s = raw.replace(" ", "")
    if ("이관" in raw) and ("폐업" in raw):
        return "transfer"
    if s == "상담" or re.match(r"^상담\d*$", s):
        return "consult"
    if s in ("소계", "합계"):
        return None
    if re.match(r"^기장\d*$", s) or (s.startswith("기장") and len(s) <= 8):
        return "bookkeeping"
    if re.match(r"^성실\d*$", s) or (s.startswith("성실") and len(s) <= 8):
        return "sincere"
    if re.match(r"^업체\d*$", s) or (s.startswith("업체") and len(s) <= 8):
        return "corp_client"
    if re.match(r"^신고\d*$", s) or (s.startswith("신고") and len(s) <= 8):
        return "agent"
    return None


def _section_dict(m: dict, start_r: int, end_r: int, id_prefix: str = "") -> dict:
    kind = m["kind"]
    sid = f"{id_prefix}{kind}-{start_r}" if id_prefix else kind
    return {
        "id": sid,
        "kind": kind,
        "label": SECTION_LABELS[kind],
        "startR": start_r,
        "endR": end_r,
        "rowCount": end_r - start_r + 1,
        "marker": m["marker"],
    }


def _build_marker_sections(markers: list[dict], end_r: int, id_prefix: str = "") -> list[dict]:
    out: list[dict] = []
    for i, m in enumerate(markers):
        start_r = m["r"]
        next_r = markers[i + 1]["r"] if i + 1 < len(markers) else end_r + 1
        sec_end = next_r - 1
        out.append(_section_dict(m, start_r, sec_end, id_prefix))
    return out


def detect_income_sections(ws_val, min_r: int, max_r: int, marker_col: int = INCOME_MARKER_COL) -> list[dict]:
    markers: list[dict] = []
    for r in range(2, max_r + 1):
        v = ws_val.cell(r, marker_col).value
        kind = classify_section_marker(v)
        if kind:
            markers.append({"r": r, "kind": kind, "marker": str(v).strip() if v else ""})

    transfer_i = next((i for i, m in enumerate(markers) if m["kind"] == "transfer"), None)
    consult_i = next((i for i, m in enumerate(markers) if m["kind"] == "consult"), None)

    split_at = transfer_i if transfer_i is not None else consult_i if consult_i is not None else len(markers)
    pre_main = markers[:split_at]

    main_markers: list[dict] = []
    pre_transfer: list[dict] = []
    seen_agent = False
    for m in pre_main:
        if m["kind"] == "agent":
            seen_agent = True
            main_markers.append(m)
        elif seen_agent and m["kind"] in ("bookkeeping", "sincere", "corp_client", "agent"):
            pre_transfer.append(m)
        else:
            main_markers.append(m)

    main_end = markers[transfer_i]["r"] - 1 if transfer_i is not None else (
        markers[consult_i]["r"] - 1 if consult_i is not None else max_r
    )

    sections: list[dict] = []
    for i, m in enumerate(main_markers):
        start_r = m["r"]
        if i + 1 < len(main_markers):
            end_r = main_markers[i + 1]["r"] - 1
        elif pre_transfer:
            end_r = pre_transfer[0]["r"] - 1
        else:
            end_r = main_end
        sections.append({**_section_dict(m, start_r, end_r), "children": []})

    if transfer_i is not None:
        child_markers = pre_transfer + markers[transfer_i + 1 : (consult_i if consult_i is not None else len(markers))]
        child_end = markers[consult_i]["r"] - 1 if consult_i is not None else max_r
        children = _build_marker_sections(child_markers, child_end, "transfer-")
        tm = markers[transfer_i]
        transfer_sec = {
            **_section_dict(tm, tm["r"], children[-1]["endR"] if children else child_end),
            "id": "transfer",
            "label": SECTION_LABELS["transfer"],
            "children": children,
        }
        sections.append(transfer_sec)

    if consult_i is not None:
        cm = markers[consult_i]
        sections.append(
            {
                **_section_dict(cm, cm["r"], max_r),
                "id": "consult",
                "label": SECTION_LABELS["consult"],
                "children": [],
            }
        )

    return sections


def has_transfer_signal(ws_val, r: int) -> bool:
    a = ws_val.cell(r, 1).value
    if isinstance(a, str) and any(k in a for k in TRANSFER_KEYWORDS):
        return True
    for c in range(2, 30):
        v = ws_val.cell(r, c).value
        if isinstance(v, str) and any(k in v for k in TRANSFER_KEYWORDS):
            return True
    return False


def is_marker_row(ws_val, r: int, marker_col: int = INCOME_MARKER_COL) -> bool:
    return classify_section_marker(ws_val.cell(r, marker_col).value) is not None


def is_data_row(ws_val, r: int, marker_col: int = INCOME_MARKER_COL) -> bool:
    if is_marker_row(ws_val, r, marker_col):
        return False
    name = ws_val.cell(r, 3).value
    no = ws_val.cell(r, 2).value
    if name in ("소계", "합계"):
        return False
    if isinstance(no, str) and no.strip() in ("소계", "합계"):
        return False
    if name is None and no is None:
        return False
    if name is None and not isinstance(no, (int, float)):
        return False
    return True


def find_row_section(r: int, sections: list[dict]) -> tuple[dict | None, dict | None]:
    for sec in sections:
        if sec.get("kind") == "transfer":
            for ch in sec.get("children") or []:
                if ch["startR"] <= r <= ch["endR"]:
                    return sec, ch
    for sec in sections:
        if sec.get("children"):
            continue
        if sec["startR"] <= r <= sec["endR"]:
            return sec, sec
    return None, None


def build_income_records(ws_val, sections: list[dict], marker_col: int = INCOME_MARKER_COL) -> list[dict]:
    records: list[dict] = []
    for r in range(2, ws_val.max_row + 1):
        if not is_data_row(ws_val, r, marker_col):
            continue
        parent, sec = find_row_section(r, sections)
        if not sec:
            continue
        transfer_flag = has_transfer_signal(ws_val, r)
        display_group = "transfer" if (sec["kind"] == "transfer" or parent and parent["kind"] == "transfer" or transfer_flag) else (
            "consult" if sec["kind"] == "consult" else "active"
        )
        rec = {
            "row": r,
            "sectionId": sec["id"],
            "sectionLabel": sec["label"],
            "parentSectionId": parent["id"] if parent and parent is not sec else sec["id"],
            "displayGroup": display_group,
            "no": ws_val.cell(r, 2).value,
            "name": ws_val.cell(r, 3).value,
            "company": ws_val.cell(r, 4).value,
            "phone": ws_val.cell(r, 5).value,
            "bizType": ws_val.cell(r, 8).value,
            "income": ws_val.cell(r, 12).value,
            "incomeTax": ws_val.cell(r, 13).value,
            "fee": ws_val.cell(r, 14).value,
            "unpaid": ws_val.cell(r, 18).value,
            "note": ws_val.cell(r, 21).value,
            "yearTag": ws_val.cell(r, 1).value if ws_val.cell(r, 1).value else None,
            "flags": [],
        }
        if transfer_flag:
            rec["flags"].append("이관/폐업")
        if rec["unpaid"] not in (None, "", 0):
            rec["flags"].append("미수")
        if rec["note"]:
            rec["flags"].append("비고")
        records.append(rec)
    return records


def load_theme_palette(xlsx_path: Path) -> list[str]:
    with zipfile.ZipFile(xlsx_path) as z:
        xml = z.read("xl/theme/theme1.xml")
    root = ET.fromstring(xml)
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    by_name: dict[str, str] = {}
    for el in root.findall(".//a:clrScheme/a:*", ns):
        name = el.tag.split("}")[-1]
        srgb = el.find("a:srgbClr", ns)
        sys_clr = el.find("a:sysClr", ns)
        if srgb is not None and srgb.get("val"):
            by_name[name] = srgb.get("val").upper()
        elif sys_clr is not None and sys_clr.get("lastClr"):
            by_name[name] = sys_clr.get("lastClr").upper()
    return [f"#{by_name.get(name, 'FFFFFF')}" for name in THEME_ORDER]


def _hex_to_rgb(hex6: str) -> tuple[int, int, int]:
    h = hex6.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _rgb_to_hex(r: int, g: int, b: int) -> str:
    return f"#{r:02X}{g:02X}{b:02X}"


def apply_tint(hex6: str, tint: float) -> str:
    r, g, b = _hex_to_rgb(hex6)
    if tint < 0:
        f = 1.0 + tint
        r, g, b = int(r * f), int(g * f), int(b * f)
    elif tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    return _rgb_to_hex(r, g, b)


def argb_to_hex(argb: str | None) -> str | None:
    if not argb:
        return None
    s = argb.upper()
    if len(s) == 8:
        rgb = s[2:]
        if rgb in ("FFFFFF", "000000") and s[:2] == "00":
            return WHITE if rgb == "FFFFFF" else BLACK
        return f"#{rgb}"
    if len(s) == 6:
        return f"#{s}"
    return None


def resolve_color(color, theme_palette: list[str]) -> str | None:
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb" and color.rgb:
        hx = argb_to_hex(color.rgb)
        return None if hx == WHITE else hx
    if ctype == "indexed" and color.indexed is not None:
        from openpyxl.styles.colors import COLOR_INDEX

        idx = int(color.indexed)
        if idx in COLOR_INDEX:
            hx = argb_to_hex(COLOR_INDEX[idx])
            return None if hx == WHITE else hx
    if ctype == "theme" and color.theme is not None:
        idx = int(color.theme)
        if 0 <= idx < len(theme_palette):
            base = theme_palette[idx].lstrip("#")
            tint = float(color.tint or 0)
            hx = apply_tint(base, tint) if tint else f"#{base}"
            return None if hx.upper() == WHITE else hx
    return None


def cell_fill_hex(cell, theme_palette: list[str]) -> str | None:
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return None
    # solid 채우기는 fgColor만 사용. end_color는 엑셀 기본값(검정)이라 무시.
    pattern = getattr(fill, "patternType", None)
    if pattern in (None, "solid"):
        attrs = ("fgColor", "start_color")
    else:
        attrs = ("fgColor", "start_color", "end_color")
    for attr in attrs:
        c = getattr(fill, attr, None)
        if not c:
            continue
        hx = resolve_color(c, theme_palette)
        if hx:
            return hx
    return None


def cell_font_style(cell, theme_palette: list[str]) -> dict:
    font = cell.font
    if not font:
        return {}
    out: dict = {}
    if font.bold:
        out["b"] = 1
    if font.italic:
        out["i"] = 1
    if font.underline and font.underline != "none":
        out["u"] = 1
    if font.color:
        fg = resolve_color(font.color, theme_palette)
        if fg and fg != BLACK:
            out["fg"] = fg
    align = cell.alignment
    if align and align.horizontal and align.horizontal != "general":
        out["ha"] = align.horizontal
    if align and align.vertical and align.vertical not in ("bottom", None):
        out["va"] = align.vertical
    return out


def serialize_value(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S") if val.time() else val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, float) and val.is_integer():
        return int(val)
    if isinstance(val, str):
        return val.strip() if val.strip() != val else val
    return val


def parse_merge(range_str: str) -> dict:
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str.upper())
    if not m:
        return {}
    from openpyxl.utils import column_index_from_string

    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    return {
        "r": r1,
        "c": column_index_from_string(c1),
        "rs": r2 - r1 + 1,
        "cs": column_index_from_string(c2) - column_index_from_string(c1) + 1,
    }


def sheet_bounds(ws_val, ws_fmt) -> tuple[int, int, int, int, set[tuple[int, int]]]:
    min_r = min_c = None
    max_r = max_c = 0
    merge_cells: set[tuple[int, int]] = set()
    for mr in ws_fmt.merged_cells.ranges:
        info = parse_merge(str(mr))
        if not info:
            continue
        r0, c0, rs, cs = info["r"], info["c"], info["rs"], info["cs"]
        for dr in range(rs):
            for dc in range(cs):
                merge_cells.add((r0 + dr, c0 + dc))
        min_r = r0 if min_r is None else min(min_r, r0)
        max_r = max(max_r, r0 + rs - 1)
        min_c = c0 if min_c is None else min(min_c, c0)
        max_c = max(max_c, c0 + cs - 1)

    for r in range(1, ws_val.max_row + 1):
        for c in range(1, ws_val.max_column + 1):
            if ws_val.cell(r, c).value is None:
                continue
            min_r = r if min_r is None else min(min_r, r)
            max_r = max(max_r, r)
            min_c = c if min_c is None else min(min_c, c)
            max_c = max(max_c, c)
    if min_r is None:
        return 1, 1, 1, 1, merge_cells
    return min_r, max_r, min_c, max_c, merge_cells


def row_fill_by_row(ws_fmt, theme_palette: list[str], min_r: int, max_r: int, anchor_col: int) -> dict[int, str]:
    """행 대표 색은 첫 데이터 열(보통 B열) 기준으로만 전파합니다."""
    out: dict[int, str] = {}
    for r in range(min_r, max_r + 1):
        bg = cell_fill_hex(ws_fmt.cell(r, anchor_col), theme_palette)
        if bg:
            out[r] = bg
    return out


def sheet_meta(name: str) -> dict:
    prefix = "종소세 25년 "
    if name.startswith(prefix):
        owner = name[len(prefix) :].strip()
        return {"kind": "income", "owner": owner}
    if name == "법인세(26.3)" or name.startswith("법인세(") and name.endswith(")"):
        return {
            "kind": "corp",
            "blocks": [
                {"owner": owner, "minC": min_c, "maxC": max_c}
                for owner, min_c, max_c in CORP_BLOCKS
            ],
        }
    if name == "법인세 조정료25":
        return {"kind": "corp_fee", "masterOnly": True}
    return {}


def export_sheet(ws_val, ws_fmt, theme_palette: list[str]) -> dict:
    min_r, max_r, min_c, max_c, merge_cells = sheet_bounds(ws_val, ws_fmt)
    row_fills = row_fill_by_row(ws_fmt, theme_palette, min_r, max_r, min_c)

    merges = []
    covered: set[tuple[int, int]] = set()
    for mr in ws_fmt.merged_cells.ranges:
        info = parse_merge(str(mr))
        if not info:
            continue
        r0, c0, rs, cs = info["r"], info["c"], info["rs"], info["cs"]
        if r0 < min_r or c0 < min_c or r0 > max_r or c0 > max_c:
            continue
        merges.append(info)
        for dr in range(rs):
            for dc in range(cs):
                if dr or dc:
                    covered.add((r0 + dr, c0 + dc))

    col_widths: dict[str, float] = {}
    for c in range(min_c, max_c + 1):
        letter = get_column_letter(c)
        dim = ws_fmt.column_dimensions.get(letter)
        if dim and dim.width:
            col_widths[letter] = round(float(dim.width), 2)

    cells = []
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            if (r, c) in covered:
                continue
            v = serialize_value(ws_val.cell(r, c).value)
            fmt_cell = ws_fmt.cell(r, c)
            bg = cell_fill_hex(fmt_cell, theme_palette)
            if not bg and r in row_fills:
                bg = row_fills[r]
            style = cell_font_style(fmt_cell, theme_palette)
            if v is None and not bg and not style:
                continue
            item: dict = {"r": r, "c": c}
            if v is not None:
                item["v"] = v
            if bg:
                item["bg"] = bg
            item.update(style)
            cells.append(item)

    return {
        "name": ws_val.title,
        "meta": sheet_meta(ws_val.title),
        "minR": min_r,
        "maxR": max_r,
        "minC": min_c,
        "maxC": max_c,
        "merges": merges,
        "colWidths": col_widths,
        "cells": cells,
    }


def attach_income_sections(sheet: dict, ws_val) -> None:
    meta = sheet.get("meta") or {}
    if meta.get("kind") != "income":
        return
    meta["markerCol"] = INCOME_MARKER_COL
    sections = detect_income_sections(ws_val, sheet["minR"], sheet["maxR"], INCOME_MARKER_COL)
    meta["sections"] = sections
    meta["records"] = build_income_records(ws_val, sections, INCOME_MARKER_COL)
    sheet["meta"] = meta


FEE_STAFF_ORDER = ["블루", "다야", "윈터", "리아", "페리", "인디"]


def detect_fee_segments_meta(ws_val, min_r: int, max_r: int) -> list[dict]:
    starts: list[int] = []
    for r in range(2, max_r + 1):
        no = ws_val.cell(r, 1).value
        company = ws_val.cell(r, 2).value
        if no == 1 and company:
            starts.append(r)
    segments: list[dict] = []
    for i, start_r in enumerate(starts):
        staff = FEE_STAFF_ORDER[i] if i < len(FEE_STAFF_ORDER) else "기타"
        end_r = (starts[i + 1] - 1) if i + 1 < len(starts) else max_r
        segments.append({"staff": staff, "startR": start_r, "endR": end_r})
    return segments


def attach_fee_sections(sheet: dict, ws_val) -> None:
    meta = sheet.get("meta") or {}
    if meta.get("kind") != "corp_fee":
        return
    meta["feeSegments"] = detect_fee_segments_meta(ws_val, sheet["minR"], sheet["maxR"])
    sheet["meta"] = meta


def import_workbook(xlsx_path: Path, out_paths: list[Path]) -> dict:
    theme_palette = load_theme_palette(xlsx_path)
    wb_val = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    wb_fmt = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=False)

    payload = {
        "version": "grid-1.0",
        "source": str(xlsx_path),
        "importedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "sheets": [],
    }
    for name in wb_val.sheetnames:
        sheet = export_sheet(wb_val[name], wb_fmt[name], theme_palette)
        attach_income_sections(sheet, wb_val[name])
        attach_fee_sections(sheet, wb_val[name])
        payload["sheets"].append(sheet)

    text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    for out in out_paths:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(text, encoding="utf-8")
    return payload


def main():
    root = Path(__file__).resolve().parents[1]
    default_xlsx = Path(r"c:\Users\찰리\Desktop\검토표.xlsx")
    parser = argparse.ArgumentParser(description="검토표.xlsx → review-grid.json")
    parser.add_argument("xlsx", nargs="?", default=str(default_xlsx))
    parser.add_argument(
        "-o",
        "--output",
        action="append",
        help="출력 JSON 경로 (기본: prototypes/assets/review-grid.json)",
    )
    args = parser.parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"파일 없음: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    outs = [Path(p) for p in args.output] if args.output else [
        root / "prototypes" / "assets" / "review-grid.json",
    ]
    payload = import_workbook(xlsx_path, outs)
    total_cells = sum(len(s["cells"]) for s in payload["sheets"])
    print(f"OK: {len(payload['sheets'])} sheets, {total_cells} cells → {outs[0]}")


if __name__ == "__main__":
    main()
